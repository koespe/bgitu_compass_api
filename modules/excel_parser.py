import datetime
import io
import re
from typing import Union

import openpyxl
from icecream import ic
from openpyxl.cell import MergedCell, Cell
from openpyxl.worksheet.worksheet import Worksheet

from data import WEEKDAY_INDEX, WEEKDAYS_LIST
from database.base import insert_schedule, manage_groups

NO_LESSONS_MAGISTRACY_STRING = "Дни самостоятельной"  # На всякий случай сократим, вдруг опечатка будет
WEEKDAY_COLUMN = 1
COLUMN_TIME_BUILDING_1 = 2
COLUMN_TIME_BUILDING_2 = 3
COLUMN_STEP_FOR_LESSON = 2  # В одной клетке хранится четное, в другой нечетное


async def process_schedule_file(file):
    """
    Инициализация расписания из excel файлов в db.groups.rawSchedule
    """
    file = io.BytesIO(file.read())
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]

    # Определяем в каком row находятся названия групп
    groups_row = None
    for row_number in range(1, 10):
        cell = parse_cell(sheet, row=row_number, col=2)
        if cell is not None and cell != "":
            if "корпус" in cell.lower():
                groups_row = row_number
                break
    ic(groups_row)

    """ Определяем клетку, где появляется понедельник. В БАК 2 клетки, в МАГ одна
    В БАК группа обычно состоит из 2х merged cell в высоту, но в нижней бывают подгруппы "а" и "б".
    С МАГ ничего не ломается, все равно попадаем на одиночную в высоту клетку"""
    schedule_start_row = None
    for row_number in range(groups_row, 20):
        cell = parse_cell(sheet, row=row_number, col=WEEKDAY_COLUMN)
        if cell:
            if "понедельник" in cell.lower():
                schedule_start_row = row_number
                groups_row = schedule_start_row - 1
                break
    ic(schedule_start_row)

    for group_column in range(4, sheet.max_column + 1):
        group_name = parse_group_name(sheet, group_column, groups_row)
        if not group_name:
            break
        group_id = await manage_groups(group_name)

        schedule_week = await parse_group_schedule(sheet, group_column, schedule_start_row)
        ic(schedule_week)
        await insert_schedule(group_id, schedule_week)


def parse_group_name(sheet, group_column, groups_row):
    group_name = parse_cell(sheet=sheet, row=groups_row, col=group_column)
    if group_name is None:
        return None

    if group_name.upper().strip() in ("А", "Б"):  # Подгруппы
        subgroup = group_name
        group_name = parse_cell(sheet, row=groups_row - 1, col=group_column).strip() + f"({subgroup})"

    group_name = group_name.strip().replace("/", "-").replace(" ", "").replace("спо", "СПО")

    ic(group_name)
    return group_name


async def parse_group_schedule(sheet, group_column, schedule_start_row):
    schedule_week = {"first_week": {}, "second_week": {}}
    schedule_per_day_first_week = []
    schedule_per_day_second_week = []

    weekday_last = parse_cell(sheet, row=schedule_start_row, col=WEEKDAY_COLUMN)
    for group_row in range(schedule_start_row, sheet.max_row + 1, COLUMN_STEP_FOR_LESSON):
        weekday_now = parse_cell(sheet, row=group_row, col=WEEKDAY_COLUMN)

        if weekday_now != weekday_last:  # Получено расписание на день, записываем его
            weekday_index = WEEKDAY_INDEX[weekday_last.lower()]
            schedule_week["first_week"][weekday_index] = schedule_per_day_first_week
            schedule_week["second_week"][weekday_index] = schedule_per_day_second_week

            if weekday_now:  # Переход на следующий день
                if weekday_now.lower() not in WEEKDAYS_LIST:  # Возможно парсер перешел на комментарий
                    break
                weekday_last = weekday_now
                schedule_per_day_first_week = []
                schedule_per_day_second_week = []
            else:  # Расписания группы закончилось, далее пустые клетки
                break

        lesson_first_week = parse_day(sheet, row=group_row, col=group_column)
        lesson_second_week = parse_day(sheet, row=group_row + 1, col=group_column)

        if lesson_first_week:
            schedule_per_day_first_week.append(lesson_first_week)
        if lesson_second_week:
            schedule_per_day_second_week.append(lesson_second_week)

    # Обрабатываем последний день недели на случай, если он не был записан
    schedule_week["first_week"][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_first_week
    schedule_week["second_week"][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_second_week

    return ensure_weekdays_keys(schedule_week)


def parse_day(sheet, row, col):
    cell = parse_cell(sheet, row=row, col=col, return_value=False)  # Ячейка со всеми ее свойствами
    cell_value: str = parse_cell(sheet, row=row, col=col, return_value=True)  # Plain text

    if (
        (cell_value is None)
        or (cell_value == "")
        or (cell_value == " ")
        or (NO_LESSONS_MAGISTRACY_STRING in str(cell_value))
    ):
        return None

    for i in range(2, 10):  # В тексте бывает много пробелов
        cell_value = cell_value.replace(" " * i, " ")
    cell_value = cell_value.strip().replace("\n", " ")

    if cell_value == "":
        return None

    # Проверка на второй корпус по цвету текста
    building = "1"
    if cell.font.color is not None:
        if cell.font.color.rgb == "FFFF0000" or cell.font.color.indexed == 10:
            building = "2"

    # Проверка на тип занятия по bold
    is_lecture = True if cell.font.bold else False

    # Следуя из корпуса получаем время
    column_time = COLUMN_TIME_BUILDING_1 if building == "1" else COLUMN_TIME_BUILDING_2
    start_at, end_at = (
        datetime.datetime.strptime(t, "%H.%M").strftime("%H:%M:%S")
        for t in parse_cell(sheet, row, column_time).split("-")
    )

    additional_data_matches = re.search(r"\[(.*?)]", cell_value)
    if additional_data_matches:
        additional_data = f" {additional_data_matches.group(1)}"
        cell_value = re.sub(r"\[.*?]", "", cell_value)  # Временное удаление этого текста
    else:
        additional_data = ""

    classroom = ""
    teacher = ""
    classrooms_count = cell_value.count("№") + cell_value.upper().count("ДОТ")
    if classrooms_count == 0:  # Преподаватель отсутствует и строчка из себя представляет название предмета
        is_lecture = False  # На всякий случай
    elif classrooms_count == 1:
        teacher = re.search(r"\((.*?)\)", cell_value).group(1)
        cell_value = re.sub(r"\([^)]*\)", "", cell_value)  # Убираем из строки преподавателя

        if "ДОТ" in cell_value.upper():
            building = "ДОТ"
            classroom = "ДОТ"
            cell_value = cell_value.replace("ДОТ", "").replace("дот", "")
        else:  # Это нормальная пара
            classroom = re.search(r"№(\S+)", cell_value).group(1)
            if "/" in classroom:  # "123/321" — разделение на первую/вторую недели
                bottom_cell_value = parse_cell(sheet, row + 1, col)
                classroom = classroom.split("/")[0 if cell.value == bottom_cell_value else 1]
            cell_value = re.sub(r"№\S+", "", cell_value).strip()  # Убираем из строки аудиторию
    elif classrooms_count == 2:
        # В этом случае первое значение — группа А, второе — Б
        teachers = [match.group(1) for match in re.finditer(r"\((.*?)\)", cell_value)]
        classrooms = [match.group(1) for match in re.finditer(r"№(\S+)", cell_value)]

        cell_value = re.sub(r"\([^)]*\)", "", cell_value)
        cell_value = re.sub(r"№\S+", "", cell_value).strip()
        cell_value = cell_value.replace(",", "")  # Между преподавателями остается запятая

        right_cell_value = parse_cell(sheet, row, col + 1)
        is_group_a = cell.value == right_cell_value
        classroom = classrooms[0 if is_group_a else 1]
        teacher = teachers[0 if is_group_a else 1]

    subject_name = cell_value.strip()
    subject_name = subject_name[0].upper() + subject_name[1:]
    subject_name += additional_data

    teacher = standardize_names(teacher)

    return {
        "subjectName": subject_name,
        "building": building,
        "startAt": start_at,
        "endAt": end_at,
        "classroom": classroom,
        "teacher": teacher,
        "isLecture": is_lecture,
    }


def parse_cell(sheet: Worksheet, row, col, using_merged=True, return_value=True) -> Union[str, MergedCell, Cell, None]:
    cell = sheet.cell(row=row, column=col)
    if using_merged:
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        else:
            cell = sheet.cell(row, col)
        if return_value:
            value = cell.value  # Plain text
        else:
            return cell
    else:
        value = cell.value
    ic(row, col, value)
    return value


def standardize_names(s):
    if s != "" and s is not None:
        s = s.replace(",", "")
        s = s.strip()

        # Замена множественных пробелов на один
        s = re.sub(r"\s+", " ", s)

        # Убираем пробелы между инициалами
        s = re.sub(r"([А-ЯЁA-Z])\.\s+([А-ЯЁA-Z]\.)", r"\1.\2", s)
    return s


def ensure_weekdays_keys(schedule_dict):
    """
    В магистратуре может не оказаться какого-то дня недели, так что добавляем их
    """
    for week in ["first_week", "second_week"]:
        for day in range(1, 7):
            if day not in schedule_dict[week]:
                schedule_dict[week][day] = []

    return schedule_dict
