import datetime
import re

import openpyxl
from icecream import ic
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_DIRECTORY
from data import WEEKDAY_INDEX
from database.base import correct_subject_spelling, manage_groups, db_init, insert_schedule, db_reset_schedules, db_drop_tables
from modules.excel_parser.excel_parser_functions import make_dict_day, extract_classrooms, standardize_names, \
    split_number_and_surname


async def initialize_excel_schedules(full_reset=False,
                                     reset_schedules=False,
                                     reset_schedules_and_subjects=False):
    """
    Инициализация расписания из excel файлов в db.groups.rawSchedule

    :param full_reset: drop tables + init tables
    :param reset_schedules: Delete Lessons
    :return:
    """
    if full_reset:
        await db_drop_tables()
        await db_init()

    if reset_schedules:
        await db_reset_schedules()

    if reset_schedules_and_subjects:
        await db_reset_schedules(and_subjects=True)

    excel_files = [file for file in EXCEL_DIRECTORY.iterdir() if '.xlsx' in file.name]

    for file in excel_files:
        wb = openpyxl.load_workbook(file)
        sheet = wb.worksheets[0]
        if 'маг' in file.name.lower():
            await process_magistracy(sheet)
        else:  # БАК и СПО
            await process_sheet(sheet)


async def process_sheet(sheet):
    """Обрабатывает лист Excel и извлекает расписание для каждой группы."""

    for group_column in range(4, sheet.max_column + 1):
        group_name = parse_group_name(sheet, group_column)
        if group_name is None:  # Пустая ячейка
            continue

        group_id = await manage_groups(group_name)

        schedule_week = await parse_group_schedule(sheet, group_id, group_column)
        await insert_schedule(group_id, schedule_week)


def parse_group_name(sheet, group_column):
    """Извлекает имя группы из листа Excel."""

    group_name = parse_cell(sheet=sheet, row=3, col=group_column)
    if group_name is None:
        return None

    group_name = group_name.upper()
    group_name = group_name.replace(' ', '')
    group_name = group_name.replace('/', '-')

    if group_name.lower() in ('а', 'б'):  # Подгруппы
        group_name = (parse_cell(sheet, row=2, col=group_column).split())[0] + f'({group_name})'

    if 'спо' in group_name.lower():
        group_name = group_name.replace('спо', 'СПО')
    ic(group_name)
    return group_name


async def find_group_information(sheet):
    """
    Находит название группы, номер строки, где начинается расписание и наличие подгрупп в файле.
    """
    building_row = 0
    group_name = ''
    is_subgroups = False

    for row in range(1, 20):
        value = parse_cell(sheet, row, col=2)
        if value is not None:
            match = re.search(r"группа\s+(\w+-\d+)", value)
            if match:
                group_name = match.group(1)

            if 'корпус' in value:
                building_row = row
                break

    if group_name == '':
        value = parse_cell(sheet, row=building_row, col=4)
        split_string = value.casefold().split(' ')
        group_index = split_string.index("группа")
        group_name = split_string[group_index + 1]
        is_subgroups = True
        building_row += 1

    groups_id = []
    if is_subgroups:
        groups_id.append(await manage_groups(f'{group_name}(А)(маг.)'))
        groups_id.append(await manage_groups(f'{group_name}(Б)(маг.)'))
    else:
        groups_id.append(await manage_groups(f'{group_name}(маг.)'))

    return building_row, group_name, is_subgroups, groups_id


async def process_magistracy(sheet):
    building_row, group_name, is_subgroups, groups_id = find_group_information(sheet)

    groups_range_in_file = 2 if is_subgroups else 1  # Есть в одном файле группы А и Б, их нужно итерировать
    for additional_index in range(groups_range_in_file):
        schedule_week = {
            "first_week": {},
            "second_week": {}
        }
        schedule_per_day_first_week = []
        schedule_per_day_second_week = []

        weekday_last = 'пятница'  # Первый день расписания всегда

        # Перебираем ячейки
        lesson_step = 2 if is_subgroups else 1  # Исключение: у строителей с группами две ячейки на урок
        for group_row in range(building_row, sheet.max_row + 1, lesson_step):
            # Идем по строчкам вниз и берем сразу на первую и вторую неделю если not is_subgroups
            weekday_now = parse_cell(sheet, row=group_row, col=1)

            if weekday_now not in ['пятница', 'суббота']:  # Еще не дошли до учебных дней
                continue
            if weekday_now == '' or weekday_now is None:
                break
            if weekday_now != weekday_last:  # Смена дней
                schedule_week['first_week'][
                    WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_first_week
                schedule_week['second_week'][
                    WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_second_week

                weekday_last = weekday_now
                schedule_per_day_first_week = []  # Данные на один день на первую неделю
                schedule_per_day_second_week = []

            if is_subgroups:
                lesson_first_week = parse_day(sheet, row=group_row, col=4 + additional_index)
                lesson_second_week = lesson_first_week
            else:
                lesson_first_week = parse_day(sheet, row=group_row, col=4)
                lesson_second_week = parse_day(sheet, row=group_row, col=5)

            if lesson_first_week is not None:
                await correct_subject_spelling(lesson_first_week.get('subjectName'), groups_id[additional_index])
                day_template = make_dict_day(data=lesson_first_week)
                schedule_per_day_first_week.append(day_template)

            if lesson_second_week is not None:
                await correct_subject_spelling(lesson_second_week.get('subjectName'), groups_id[additional_index])
                day_template = make_dict_day(data=lesson_second_week)
                schedule_per_day_second_week.append(day_template)
        # Закончили парсить группу и заносим последние данные
        schedule_week['first_week'][
            WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_first_week
        schedule_week['second_week'][
            WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_second_week
        await insert_schedule(groups_id[additional_index], schedule_week)


async def parse_group_schedule(sheet, group_id, group_column):
    """Парсит расписание для одной группы."""

    schedule_week = {"first_week": {}, "second_week": {}}
    schedule_per_day_first_week = []
    schedule_per_day_second_week = []

    weekday_last = parse_cell(sheet, row=4, col=1)
    lesson_step = 2

    for group_row in range(4, sheet.max_row + 1, lesson_step):
        weekday_now = parse_cell(sheet, row=group_row, col=1)

        if weekday_now != weekday_last:
            schedule_week['first_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_first_week
            schedule_week['second_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_second_week

            if weekday_now:
                weekday_last = weekday_now
                schedule_per_day_first_week = []
                schedule_per_day_second_week = []

        lesson_first_week = parse_day(sheet, row=group_row, col=group_column)
        lesson_second_week = parse_day(sheet, row=group_row + 1, col=group_column)

        if lesson_first_week:
            await correct_subject_spelling(lesson_first_week.get('subjectName'), group_id)
            day_template = make_dict_day(data=lesson_first_week)
            schedule_per_day_first_week.append(day_template)

        if lesson_second_week:
            await correct_subject_spelling(lesson_second_week.get('subjectName'), group_id)
            day_template = make_dict_day(data=lesson_second_week)
            schedule_per_day_second_week.append(day_template)

    schedule_week['first_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_first_week
    schedule_week['second_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_second_week

    # ТИПО ФИКС
    schedule_week = {"first_week": schedule_week["second_week"], "second_week": schedule_week["first_week"]}
    return schedule_week


def parse_cell(sheet: Worksheet, row, col, using_merged=True):
    cell = sheet.cell(row=row, column=col)
    if using_merged:
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        else:
            cell = sheet.cell(row, col)
        value = cell.value  # Plain text
    else:
        value = cell.value
    return value


col_weekday = 1
col_time_1_building = 2
col_time_2_building = 3


def parse_day(sheet: Worksheet, row, col):
    """
    Парсинг в пределах одной пары
    :param sheet: Таблица из файла
    :param row: Строка
    :param col: Столбец
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    else:
        cell = sheet.cell(row, col)
    value = cell.value  # Текст ячейки

    if (value is None) or (value == '') or (value == ' '):  # Обработка пустоты
        return None

    # Проверка на второй корпус по цвету текста
    building = 1
    if cell.font.color is not None:
        if cell.font.color.rgb == 'FFFF0000' or cell.font.color.indexed == 10:
            building = 2

    # Проверка на тип занятия по bold
    if cell.font.bold:
        is_lecture = True
    else:
        is_lecture = False

    # Следуя из корпуса получаем время
    if building == 2:
        lesson_time = parse_cell(sheet, row, col_time_2_building)
    else:
        lesson_time = parse_cell(sheet, row, col_time_1_building)

    try:
        lesson_time = lesson_time.split('-')
        start_at = datetime.datetime.strptime(lesson_time[0], '%H.%M').time()
        end_at = datetime.datetime.strptime(lesson_time[1], '%H.%M').time()
    except (ValueError, AttributeError) as e:  # Какая-то некорректная ячейка, не смог отловить в свое время ошибку
        return None

    value: str = value.replace("\n", " ")  # Ненужные переносы

    # Обработка фраз "консп." и временное удаления для корректной работы алгоритмы
    strange_phrase = ''
    if 'конс' in value:
        if '(' in value and ')' in value:
            strange_phrase = value[value.find('('):value.find(')') + 1]
            value = value.replace(strange_phrase, '')

    # В строчках бывает очень много пробелов
    for i in range(2, 30):
        value = value.replace(' ' * i, ' ')

    # Основа алгоритма для вытаскивания данных из ячейки
    str_list = value.split(' ')
    str_list = list(filter(None, str_list))  # Убираем '' из списка
    str_list_classrooms = str_list.copy()  # Сохранение оригинала для логики преподавателей
    if len(str_list) == 0:  # В некоторых ячейках есть непонятный невидимый символ переноса
        return None
    ic('до обработки:, ', str_list)
    str_list = split_number_and_surname(str_list)
    ic(str_list)
    classrooms_algorithm = extract_classrooms(str_list)
    classrooms = classrooms_algorithm['classrooms']
    str_list = classrooms_algorithm['str_list_no_classrooms']
    # print('classrooms: ',classrooms)

    if bool(classrooms):  # Самая обычная ячейка
        if len(classrooms) == 1:
            classroom = classrooms[0]

            """ Обработка спец символов:
            / — разделение на чет/нечет - определяем по клетке
            : или ; - просто второй кабинет, ничего не делаем """
            if '/' in classroom:
                lesson_time = parse_cell(sheet, row, col_time_1_building)
                next_lesson_time = parse_cell(sheet, row + 1, col_time_1_building)
                if lesson_time == next_lesson_time:  # Значит это первая неделя, возвращаем число перед тире
                    classroom = classroom.split('/')[0] + f'({classroom.split("/")[1]})'
                else:
                    classroom = classroom.split('/')[1] + f'({classroom.split("/")[0]})'

            # Странное написание преподавателя — Ситуация вида "СальниковаН.А."
            if str_list_classrooms[-2].replace('/', '').replace(':', '').replace(';', '').isdigit():
                teacher = str_list[-1]
                del str_list[-1:]
            else:
                # print(str_list)
                if len(str_list) >= 3:  # бывает "информатика 218" — не понятно куда пропал препод

                    # Обработка вида Луцевич А А
                    if len(str_list[-1]) == 1 and len(str_list[-2]) == 1:
                        str_list.append(str_list[-2] + '.' + str_list[-1] + '.')
                        del str_list[-3]
                        del str_list[-2]

                    teacher = str_list[-2] + ' ' + str_list[-1]
                    del str_list[-2:]
                else:
                    teacher = ''

        else:  # Иногда бывает два кабинета в одной ячейке. !!! В этом случае первое значение — группа А, второе — Б
            right_cell_value = classrooms[1] if len(classrooms) > 1 else classrooms[0]
            ic('ТРИГГЕР 2 ПРЕПОДА')
            right_cell_value = parse_cell(sheet, row, col + 1)
            if cell.value == right_cell_value:
                classroom = classrooms[0]
                teacher = str_list[-4] + ' ' + str_list[-3]

            else:
                classroom = classrooms[1]
                teacher = str_list[-2] + ' ' + str_list[-1]
            del str_list[-4:]  # Удаляем оба преподавателя

        ic('в standartize идет: ', teacher)
        # ic('После очистки массива от препода: ',str_list)
        teacher = standardize_names(teacher)  # Правим написание преподавателя
        ic(teacher)
        subject_name = ' '.join(str_list)
        if classroom.upper() == 'ДОТ':
            building = 'ДОТ'
            classroom = 'ДОТ'

    else:  # Преподаватель отсутствует и строчка из себя представляет название предмета
        subject_name = ' '.join(str_list)
        classroom = ''
        teacher = ''
        is_lecture = False

    if len(strange_phrase) > 0:
        subject_name += ' ' + strange_phrase
    subject_name = ' '.join(str_list)
    subject_name = subject_name[0].upper() + subject_name[1:]

    return {
        "subjectName": subject_name,
        "building": str(building),
        "startAt": start_at,
        "endAt": end_at,
        "classroom": classroom,
        "teacher": teacher,
        "isLecture": is_lecture
    }
